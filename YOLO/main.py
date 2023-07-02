import cv2
import numpy as np
import openpyxl
from openpyxl.chart import BarChart, Reference


import time
import sys
import os
import glob

from openpyxl.workbook import Workbook

CONFIDENCE = 0.5
SCORE_THRESHOLD = 0.5
IOU_THRESHOLD = 0.5
data = []
unique_data = []
summary = []
sum_dogs = 0
sum_cats = 0

# конфиг YOLO
config_path = "venv/cfg/yolov3.cfg"
# файл с весами
weights_path = "venv/weights/yolov3.weights"


# загрузка всех меток классов (объектов)
labels = open("venv/data/coco.names").read().strip().split("\n")
# генерируем цвета для каждого объекта и последующего построения
colors = np.random.randint(0, 255, size=(len(labels), 3), dtype="uint8")

# загрузка модели
net = cv2.dnn.readNetFromDarknet(config_path, weights_path)

# Путь к папке
path = "venv/images"
extension = "jpg"
os.chdir(path)

# Перебор всех файлов в папке
for file in glob.glob('*.' + extension):
    # Загрузка картинок
    # path_name = "venv/images/dva.jpg"
    path_name = os.path.basename(file)  # Путь
    image = cv2.imread(path_name)   # Взятие картинки для обработки
    file_name = os.path.basename(path_name) # Имя файла
    filename, ext = file_name.split(".")

    # нормализация значения пикселей в диапазоне от 0 до 1
    # и изменение размера изображения до (416, 416)
    h, w = image.shape[:2] # Берем ширину и высоту
    # создаем 4D blob(переводим изображение в последовательность байтов)
    blob = cv2.dnn.blobFromImage(image, 1/255.0, (416, 416), swapRB=True, crop=False)

    # устанавливает blob как вход сети
    net.setInput(blob)
    # получаем имена всех слоев
    ln = net.getLayerNames()
    try:
        ln = [ln[i[0] - 1] for i in net.getUnconnectedOutLayers()]
    except IndexError:
        # в случае, если CUDA недоступен
        ln = [ln[i - 1] for i in net.getUnconnectedOutLayers()]

    # прямая связь (вывод) и получение выхода сети
    # измерение времени для обработки в секундах
    start = time.perf_counter()
    layer_outputs = net.forward(ln)
    time_took = time.perf_counter() - start
    print(f"Потребовалось: {time_took:.2f}s")

    font_scale = 1
    thickness = 1
    boxes, confidences, class_ids = [], [], []
    # перебираем каждый из выходов слоя
    for output in layer_outputs:
        # перебираем каждое обнаружение объекта
        for detection in output:
            # извлекаем идентификатор класса (метку) и достоверность (как вероятность)
            # обнаружение текущего объекта
            scores = detection[5:]
            class_id = np.argmax(scores)
            confidence = scores[class_id]
            # отбрасываем слабые прогнозы, убедившись, что обнаруженная
            # вероятность больше минимальной вероятности
            if confidence > CONFIDENCE:
                # масштабируем координаты ограничивающего прямоугольника относительно
                # размера изображения, учитывая, что YOLO
                # возвращает центральные координаты (x, y) ограничивающего
                # поля, за которым следуют ширина и высота поля
                box = detection[:4] * np.array([w, h, w, h])
                (centerX, centerY, width, height) = box.astype("int")
                # используем центральные координаты (x, y) для получения вершины
                # и левый угол ограничительной рамки
                x = int(centerX - (width / 2))
                y = int(centerY - (height / 2))
                # обновляем наш список координат ограничивающего прямоугольника, достоверности,
                # и идентификаторы класса
                boxes.append([x, y, int(width), int(height)])
                confidences.append(float(confidence))
                class_ids.append(class_id)

    # выполнение не максимального подавления с учетом оценок, определенных ранее
    idxs = cv2.dnn.NMSBoxes(boxes, confidences, SCORE_THRESHOLD, IOU_THRESHOLD)

    # Проверка на обнаружение хотя бы одиного объекта
    if len(idxs) > 0:
        # Перебираем сохраняемые индексы
        for i in idxs.flatten():
            # извлекаем координаты ограничивающего прямоугольника
            x, y = boxes[i][0], boxes[i][1]
            w, h = boxes[i][2], boxes[i][3]
            # рисуем прямоугольник ограничивающей рамки и подписываем на изображении
            color = [int(c) for c in colors[class_ids[i]]]
            cv2.rectangle(image, (x, y), (x + w, y + h), color=color, thickness=thickness)
            text = f"{labels[class_ids[i]]}: {confidences[i]:.2f}"

            #List найденных лейблов
            data.append(labels[class_ids[i]])
            unique_data = list(set(data))

            #List найденных лейблов

            # вычисляем ширину и высоту текста, чтобы определить прозрачные поля в качестве фона текста
            (text_width, text_height) = cv2.getTextSize(text, cv2.FONT_HERSHEY_SIMPLEX, fontScale=font_scale, thickness=thickness)[0]
            text_offset_x = x
            text_offset_y = y - 5
            box_coords = ((text_offset_x, text_offset_y), (text_offset_x + text_width + 2, text_offset_y - text_height))
            overlay = image.copy()
            cv2.rectangle(overlay, box_coords[0], box_coords[1], color=color, thickness=cv2.FILLED)
            # добавляем непрозрачность (прозрачность поля)
            image = cv2.addWeighted(overlay, 0.6, image, 0.4, 0)
            # Помечаем текст (метка: доверие%)
            cv2.putText(image, text, (x, y - 5), cv2.FONT_HERSHEY_SIMPLEX,
                fontScale=font_scale, color=(0, 0, 0), thickness=thickness)

    cv2.imwrite(filename + "_yolo3." + ext, image)


# Выводим статистику в Excel

sum_classes = 0

for j in range(0, len(unique_data)):
    for i in range(0, len(data)):
        if unique_data[j] == data[i]:
            sum_classes = sum_classes + 1
    summary.append(sum_classes)
    sum_classes = 0

print(unique_data)
print(summary)

# Создаем новый документ Excel
workbook = Workbook()
sheet = workbook.active

# Заполняем заголовки
sheet['A1'] = 'Class'
sheet['B1'] = 'Summary'

# Заполняем данные
for row in range(2, len(unique_data) + 2):
    sheet.cell(row=row, column=1).value = unique_data[row - 2]
    sheet.cell(row=row, column=2).value = summary[row - 2]

# Создаем объекты для построения диаграммы
# Объект values, который представляет собой ссылку на ячейки листа sheet в столбце 2
values = Reference(sheet, min_col=2, min_row=2, max_row=len(unique_data) + 1)
categories = Reference(sheet, min_col=1, min_row=2, max_row=len(unique_data) + 1)
chart = BarChart()
chart.title = 'Class summary'
chart.y_axis.title = 'Summary'
chart.x_axis.title = 'Class'
chart.add_data(values)
chart.set_categories(categories)

# Добавляем диаграмму на лист
sheet.add_chart(chart, "D2")

# Сохраняем документ
workbook.save("Class_statistics.xlsx")
